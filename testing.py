import re
from openpyxl import Workbook, load_workbook

# find matches where capital or non-capital letter and some characters are
# are followed by a '.' then a ' ' and then a capital/non capital mutliple
# characters and then another space or not(maybe a period or other punctuation)
# at end of sentence. will find Mr. taylor, mrs. taylor, ms. taylor ect.
regex = r"[A-Za-z]*\w[.] [A-Za-z]*\w\s?"

wb = load_workbook('feedback.xlsx')

ws = wb.active

i = 1 #set initial row to begin at

for r in ws['A']: #create tuple of values in column 'A'
    ws.cell(row=i, column = 1, value = re.sub(regex,'',r.value)) #replace teacher name with ''
    i += 1 #increase row number to move to next row

wb.save('updated.xlsx')
